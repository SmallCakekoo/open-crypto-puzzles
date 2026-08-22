# -*- coding: utf-8 -*-
import json, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = r"C:\Users\patat\Desktop\open-crypto-puzzles"
os.chdir(ROOT)

TYPE_ES = {
    "text-cipher": "cifrado de texto",
    "pixel-code": "codigo en pixeles/imagen",
    "web-tree": "arbol de paginas web",
    "raw-private-key": "clave privada en crudo",
    "bip39-seed": "semilla BIP39",
    "word-selection": "seleccion de palabras",
    "bip38": "BIP38 (clave cifrada)",
    "physical-object": "objeto fisico",
    "shamir": "reparto de secreto Shamir",
    "image-stego": "esteganografia en imagen",
    "hash-collision": "colision de hash",
    "smart-contract": "contrato inteligente",
    "timelock": "timelock (bloqueo temporal)",
    "video-series": "serie de videos",
    "book": "libro / texto publicado",
    "brainwallet": "brainwallet (frase -> clave)",
    "password-pages": "paginas con contrasena",
    "geometry": "geometria / coordenadas",
    "multisig": "multisig",
    "audio": "audio",
}

DIFF_ES = {
    "none": "Ninguna (ya resuelto)",
    "human-action": "Accion humana (solo reclamar/ejecutar)",
    "external-info": "Falta informacion externa",
    "bounded-compute": "Computo acotado (fuerza bruta viable)",
    "insight": "Requiere un golpe de ingenio (insight)",
    "uneconomic": "Computo posible pero antieconomico",
    "research-breakthrough": "Requiere avance de investigacion criptografica",
}

# Puntaje base de dificultad REAL (1-10), asignado por criterio criptologico,
# deliberadamente independiente del valor del premio.
DIFF_SCORE = {
    "none": 0,
    "human-action": 1,
    "external-info": 4,
    "bounded-compute": 5,
    "insight": 7,
    "uneconomic": 8,
    "research-breakthrough": 10,
}

STATUS_ES = {
    "open": "Abierto",
    "solved": "Resuelto",
    "dead-end": "Callejon sin salida",
    "watch": "En vigilancia",
}

TIER_ES = {
    "1-big-prizes": "1 - Premio grande",
    "2-mid-prizes": "2 - Premio medio",
    "3-small-prizes": "3 - Premio pequeno",
    "4-solved": "4 - Resueltos",
    "archive": "Archivo - Callejones sin salida",
}

def fmt_amount(prize):
    amt = prize.get("amount")
    asset = prize.get("asset", "")
    if amt is None:
        return ""
    if isinstance(amt, float):
        s = f"{amt:,.8f}".rstrip("0").rstrip(".")
    else:
        s = f"{amt:,}"
    return f"{s} {asset}"

rows = []
files = sorted(glob.glob("**/puzzle.json", recursive=True))
for f in files:
    f = f.replace(os.sep, "/")
    d = json.load(open(f, encoding="utf-8"))
    top_folder = f.split("/")[0]
    prize = d.get("prize", {}) or {}
    usd = prize.get("usd_estimate")
    types = d.get("puzzle_type", []) or []
    types_es = ", ".join(TYPE_ES.get(t, t) for t in types)
    diff_left = d.get("difficulty_left") or "none"
    leads = d.get("leads", []) or []
    top_lead = leads[0]["title"] if leads else ""
    tested = (d.get("tested_total") or {}).get("candidates")
    summary = (d.get("derivation") or {}).get("summary", "")
    note = d.get("difficulty_note", "")
    status = d.get("status", "")

    rows.append({
        "tier_folder": top_folder,
        "tier": TIER_ES.get(top_folder, top_folder),
        "titulo": d.get("title", ""),
        "estado": STATUS_ES.get(status, status),
        "cadena": d.get("chain", ""),
        "premio_nativo": fmt_amount(prize),
        "premio_usd": usd if usd is not None else "",
        "tipo_mecanismo": types_es,
        "n_etapas": len(types),
        "de_que_trata": summary,
        "dificultad_restante": DIFF_ES.get(diff_left, diff_left),
        "puntaje_dificultad": DIFF_SCORE.get(diff_left, 3),
        "candidatos_probados_sin_exito": tested if tested else 0,
        "siguiente_pista": top_lead,
        "detalle_dificultad": note,
        "slug": d.get("slug", ""),
    })

# ---- ranking cruzado dificultad vs precio, para desmentir la correlacion ----
by_usd = sorted([r for r in rows if isinstance(r["premio_usd"], (int, float))],
                key=lambda r: r["premio_usd"], reverse=True)
for i, r in enumerate(by_usd, start=1):
    r["rank_precio"] = i
by_diff = sorted(rows, key=lambda r: r["puntaje_dificultad"], reverse=True)
for i, r in enumerate(by_diff, start=1):
    r["rank_dificultad"] = i
for r in rows:
    r.setdefault("rank_precio", "")
    r.setdefault("rank_dificultad", "")

# orden final de presentacion: por tier de premio y luego por dificultad ascendente
tier_order = {"1-big-prizes": 0, "2-mid-prizes": 1, "3-small-prizes": 2, "4-solved": 3, "archive": 4}
rows.sort(key=lambda r: (tier_order.get(r["tier_folder"], 9), r["puntaje_dificultad"]))

# ================= Workbook =================
wb = Workbook()

# ---------- Hoja 1: Metodologia ----------
ws0 = wb.active
ws0.title = "Metodologia"
ws0.sheet_view.showGridLines = False
title_font = Font(size=16, bold=True, color="1F2933")
h_font = Font(size=12, bold=True, color="1F2933")
body_font = Font(size=11, color="333333")
ws0.column_dimensions["A"].width = 100
ws0["A1"] = "Como se clasifico la dificultad de cada puzzle"
ws0["A1"].font = title_font
lines = [
    "",
    "Objetivo: separar la dificultad REAL de resolver cada acertijo del tamano de su premio.",
    "El precio de un puzzle no dice nada fiable sobre lo dificil que es: puzzles baratos llevan",
    "anos sin resolverse (falta una sola pista humana) y puzzles carisimos tienen la cadena",
    "criptografica entera resuelta salvo un ultimo insight. Por eso esta hoja NO ordena por precio.",
    "",
    "Columna 'Dificultad restante' (categoria del propio analisis tecnico del repo, campo",
    "difficulty_left de cada puzzle.json):",
    "  - Ninguna: ya resuelto.",
    "  - Accion humana: la solucion se conoce, solo falta reclamarla o ejecutar un paso trivial.",
    "  - Falta informacion externa: falta un dato (una foto, una publicacion, un hint) que no",
    "    depende de razonar mas sino de que alguien lo aporte o aparezca.",
    "  - Computo acotado: la busqueda es computacionalmente viable (fuerza bruta con recursos",
    "    razonables) pero aun no se ha completado o no ha dado con la clave.",
    "  - Insight: la cadena tecnica esta resuelta pero falta una idea, lectura o interpretacion",
    "    creativa que ningun barrido por fuerza bruta reemplaza.",
    "  - Antieconomico: el computo necesario es tecnicamente posible pero cuesta mas que el",
    "    propio premio (no compensa intentarlo salvo por deporte).",
    "  - Avance de investigacion: requiere una ruptura criptografica que hoy no existe",
    "    publicamente (ej. una colision de hash real).",
    "",
    "Columna 'Puntaje de dificultad (0-10)': mapeo numerico de la categoria anterior, fijo y",
    "definido antes de mirar ningun precio:",
    "  ninguna=0, accion humana=1, falta info externa=4, computo acotado=5, insight=7,",
    "  antieconomico=8, avance de investigacion=10.",
    "",
    "Columna 'N. etapas del mecanismo': cantidad de tecnicas encadenadas (estego + BIP39 + cifrado,",
    "etc.). Es una medida de COMPLEJIDAD DE CONSTRUCCION, no de dificultad; puzzles con muchas",
    "etapas pueden estar totalmente resueltos salvo el ultimo eslabon, y puzzles de una sola",
    "etapa pueden ser irresolubles (ej. colision de hash).",
    "",
    "Columna 'Candidatos ya probados sin exito': volumen de busqueda ya descartado. Un numero alto",
    "aqui junto con 'insight' como dificultad restante es la senal mas clara de que el problema no",
    "es de fuerza bruta sino de una idea que falta.",
    "",
    "Columnas 'Rank precio' y 'Rank dificultad': puesto de cada puzzle si se ordenara por premio en",
    "USD, y puesto si se ordenara por puntaje de dificultad. Compara ambas columnas en la hoja",
    "'Puzzles': veras que NO coinciden. Ejemplos notables:",
    "  - Peter Todd (colision de hash, ~37 mil USD): dificultad maxima (10), pero premio medio.",
    "  - Guntis Vitolins (bounded-compute, ~16 mil USD): mas barato que GSMG.io (insight, 5 BTC)",
    "    y sin embargo requiere mas computo bruto ya invertido.",
    "  - Exitonly Challenge 14 (30 mil sats, ~19 USD): catalogado 'antieconomico', de las dificultades",
    "    mas altas de la lista pese a ser de los premios mas bajos.",
    "  - Keir Finlow-Bates (600 mil sats): 'accion humana', de las dificultades mas bajas, en un",
    "    premio mediano.",
    "",
    "Conclusion para tu busqueda: no asumas que un puzzle barato es facil ni que uno caro es dificil.",
    "Mira la columna de dificultad restante y la de candidatos ya probados: ahi esta la senal real.",
]
r = 2
for ln in lines:
    ws0.cell(row=r, column=1, value=ln).font = body_font
    ws0.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------- Hoja 2: Puzzles ----------
ws = wb.create_sheet("Puzzles")
ws.sheet_view.showGridLines = False

headers = [
    "Categoria de premio", "Titulo", "Estado", "Cadena",
    "Premio (nativo)", "Premio (USD est.)",
    "Mecanismo(s)", "N. etapas del mecanismo",
    "De que trata (resumen tecnico)",
    "Dificultad restante", "Puntaje de dificultad (0-10)",
    "Candidatos ya probados sin exito",
    "Siguiente pista recomendada",
    "Detalle de la dificultad",
    "Rank precio (1=mas caro)", "Rank dificultad (1=mas dificil)",
    "slug",
]
ws.append(headers)
header_fill = PatternFill("solid", fgColor="2A3140")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.freeze_panes = "A2"

diff_color = {
    0: "D1FAE5", 1: "D1FAE5",
    4: "FEF3C7", 5: "FEF3C7",
    7: "FED7AA", 8: "FED7AA",
    10: "FCA5A5",
}

for r_i, row in enumerate(rows, start=2):
    values = [
        row["tier"], row["titulo"], row["estado"], row["cadena"],
        row["premio_nativo"], row["premio_usd"],
        row["tipo_mecanismo"], row["n_etapas"],
        row["de_que_trata"],
        row["dificultad_restante"], row["puntaje_dificultad"],
        row["candidatos_probados_sin_exito"],
        row["siguiente_pista"], row["detalle_dificultad"],
        row["rank_precio"], row["rank_dificultad"],
        row["slug"],
    ]
    ws.append(values)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r_i, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 7, 9, 13, 14)))
    score_cell = ws.cell(row=r_i, column=11)
    fill_color = diff_color.get(row["puntaje_dificultad"], "E5E7EB")
    score_cell.fill = PatternFill("solid", fgColor=fill_color)
    score_cell.alignment = Alignment(horizontal="center", vertical="top")
    usd_cell = ws.cell(row=r_i, column=6)
    usd_cell.number_format = "#,##0"

widths = {
    1: 16, 2: 30, 3: 12, 4: 10, 5: 20, 6: 14, 7: 26, 8: 10,
    9: 55, 10: 26, 11: 12, 12: 16, 13: 40, 14: 60, 15: 10, 16: 12, 17: 30,
}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 34

last_row = len(rows) + 1
tab = Table(displayName="Puzzles", ref=f"A1:Q{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)

# ---------- Hoja 3: Resumen ----------
ws2 = wb.create_sheet("Resumen")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "Resumen rapido"
ws2["A1"].font = title_font
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 14

from collections import Counter
diff_counts = Counter(r["dificultad_restante"] for r in rows if r["estado"] == "Abierto")
ws2["A3"] = "Puzzles abiertos por dificultad restante"
ws2["A3"].font = h_font
rr = 4
for k, v in sorted(diff_counts.items(), key=lambda kv: -kv[1]):
    ws2.cell(row=rr, column=1, value=k)
    ws2.cell(row=rr, column=2, value=v)
    rr += 1

rr += 1
ws2.cell(row=rr, column=1, value="Puzzles abiertos mas baratos con dificultad alta (insight/uneconomico/investigacion)").font = h_font
rr += 1
candidates = [r for r in rows if r["estado"] == "Abierto" and r["puntaje_dificultad"] >= 7 and isinstance(r["premio_usd"], (int, float))]
candidates.sort(key=lambda r: r["premio_usd"])
ws2.cell(row=rr, column=1, value="Titulo")
ws2.cell(row=rr, column=2, value="USD")
ws2.cell(row=rr, column=3, value="Dificultad")
for c in range(1, 4):
    ws2.cell(row=rr, column=c).font = Font(bold=True)
rr += 1
for r in candidates[:8]:
    ws2.cell(row=rr, column=1, value=r["titulo"])
    ws2.cell(row=rr, column=2, value=r["premio_usd"])
    ws2.cell(row=rr, column=3, value=r["dificultad_restante"])
    rr += 1

out_path = os.path.join(ROOT, "puzzles_clasificados.xlsx")
wb.save(out_path)
print("OK ->", out_path, "filas:", len(rows))
